"""
report.py — Excel Report Generator
=====================================
Membuat laporan Excel hasil blast menggunakan openpyxl.
Fitur: header merah, conditional formatting, auto-width kolom.

Author  : CRM Team - PT Telkomsel Branch Karawang
Project : CRM Blast IndiHome
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from config import REPORT_FILE, RESULT_FAILED, RESULT_SKIPPED, RESULT_SUCCESS
from logger import log


# ──────────────────────────────────────────────
# Warna (Hex RGB tanpa #)
# ──────────────────────────────────────────────
COLOR_RED_HEADER = "CC0000"   # Merah Telkomsel
COLOR_WHITE      = "FFFFFF"
COLOR_SUCCESS    = "D6F5D6"   # Hijau muda
COLOR_FAILED     = "FFD6D6"   # Merah muda
COLOR_SKIPPED    = "FFF9C4"   # Kuning muda
COLOR_HEADER_TXT = "FFFFFF"   # Teks header: putih
COLOR_BORDER     = "CCCCCC"   # Border abu


def generate_report(
    results: list[dict[str, Any]],
    output_path: Path | None = None,
) -> Path:
    """
    Membuat laporan Excel dari hasil blast.

    Args:
        results     : List hasil blast [{nomor_indihome, nomor_wa, nama, status, ...}]
        output_path : Path output file (default: REPORT_FILE dari config)

    Returns:
        Path: Path file Excel yang dihasilkan

    Raises:
        Exception: Jika gagal membuat file

    Example:
        >>> results = engine.get_results()
        >>> path = generate_report(results)
        >>> print(path)
        reports/laporan_blast.xlsx
    """
    if output_path is None:
        # Tambah timestamp di nama file
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = REPORT_FILE.parent / f"laporan_blast_{ts}.xlsx"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    log.info(f"Membuat laporan Excel: {output_path}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Blast"

    # ── Header Row ───────────────────────────────────────────
    headers = [
        "No.",
        "Nomor IndiHome",
        "Nomor WA",
        "Nama",
        "Status",
        "Jam Kirim",
        "Percobaan",
        "Keterangan",
    ]

    _write_header_row(ws, headers)

    # ── Data Rows ────────────────────────────────────────────
    for row_num, result in enumerate(results, start=2):
        status = result.get("status", RESULT_FAILED)
        fill   = _get_status_fill(status)

        row_data = [
            result.get("num",            row_num - 1),
            result.get("nomor_indihome", ""),
            result.get("nomor_wa",       ""),
            result.get("nama",           ""),
            status,
            result.get("jam_kirim",      ""),
            result.get("percobaan",      1),
            result.get("keterangan",     ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border    = _thin_border()

            # Khusus kolom Status: bold dan warna teks
            if col_idx == 5:
                cell.font = Font(
                    bold=True,
                    color=_get_status_font_color(status),
                )

    # ── Summary Sheet ────────────────────────────────────────
    _add_summary_sheet(wb, results)

    # ── Auto Width ───────────────────────────────────────────
    _auto_width(ws)

    # ── Freeze Header ────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Simpan ───────────────────────────────────────────────
    wb.save(output_path)
    log.info(f"Laporan berhasil disimpan: {output_path}")
    return output_path


def _write_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, headers: list[str]) -> None:
    """Menulis baris header dengan style merah Telkomsel."""
    header_fill = PatternFill(fill_type="solid", fgColor=COLOR_RED_HEADER)
    header_font = Font(bold=True, color=COLOR_HEADER_TXT, size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = _thin_border()

    # Tinggi header row
    ws.row_dimensions[1].height = 25


def _get_status_fill(status: str) -> PatternFill:
    """Mengembalikan warna fill berdasarkan status."""
    color_map = {
        RESULT_SUCCESS: COLOR_SUCCESS,
        RESULT_FAILED : COLOR_FAILED,
        RESULT_SKIPPED: COLOR_SKIPPED,
    }
    color = color_map.get(status, COLOR_WHITE)
    return PatternFill(fill_type="solid", fgColor=color)


def _get_status_font_color(status: str) -> str:
    """Mengembalikan warna teks berdasarkan status."""
    color_map = {
        RESULT_SUCCESS: "1A7F1A",  # Hijau tua
        RESULT_FAILED : "CC0000",  # Merah
        RESULT_SKIPPED: "8B6914",  # Kuning tua
    }
    return color_map.get(status, "000000")


def _thin_border() -> Border:
    """Mengembalikan border tipis abu-abu."""
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _auto_width(ws) -> None:
    """Menyesuaikan lebar kolom secara otomatis berdasarkan konten."""
    col_widths = {
        1: 6,    # No.
        2: 18,   # Nomor IndiHome
        3: 18,   # Nomor WA
        4: 22,   # Nama
        5: 12,   # Status
        6: 12,   # Jam Kirim
        7: 12,   # Percobaan
        8: 35,   # Keterangan
    }

    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


def _add_summary_sheet(wb: openpyxl.Workbook, results: list[dict]) -> None:
    """Menambahkan sheet ringkasan statistik."""
    ws_sum = wb.create_sheet(title="Ringkasan")

    total   = len(results)
    success = sum(1 for r in results if r.get("status") == RESULT_SUCCESS)
    failed  = sum(1 for r in results if r.get("status") == RESULT_FAILED)
    skipped = sum(1 for r in results if r.get("status") == RESULT_SKIPPED)
    pct     = round(success / total * 100, 1) if total > 0 else 0

    red_fill   = PatternFill(fill_type="solid", fgColor=COLOR_RED_HEADER)
    bold_white = Font(bold=True, color="FFFFFF", size=12)

    summary_data = [
        ("Laporan CRM Blast IndiHome", ""),
        ("Tanggal", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("", ""),
        ("Total Target",  total),
        ("Berhasil",      success),
        ("Gagal",         failed),
        ("Dilewati",      skipped),
        ("Success Rate",  f"{pct}%"),
    ]

    for row_idx, (label, value) in enumerate(summary_data, start=1):
        cell_label = ws_sum.cell(row=row_idx, column=1, value=label)
        cell_value = ws_sum.cell(row=row_idx, column=2, value=value)

        if row_idx == 1:
            cell_label.fill = red_fill
            cell_label.font = bold_white
            ws_sum.merge_cells(f"A1:B1")
            cell_label.alignment = Alignment(horizontal="center")
        elif label in ("Berhasil",):
            cell_value.font = Font(bold=True, color="1A7F1A")
        elif label in ("Gagal",):
            cell_value.font = Font(bold=True, color="CC0000")

        # Hanya set font bold jika belum di-set sebelumnya (row selain header & status)
        if row_idx != 1 and label not in ("Berhasil", "Gagal"):
            cell_label.font = Font(bold=True)

    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 30
